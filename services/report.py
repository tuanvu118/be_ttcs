import io
import pandas as pd
from datetime import timezone, datetime
from typing import List, Optional

from beanie import PydanticObjectId

from exceptions import ErrorCode, app_exception
from models.report import InternalEvent,Report
from models.unit_event import UnitEvent
from models.unit_event_submissions import UnitEventSubmission, UnitEventSubmissionStatus
from repositories.report_repo import ReportRepository
from repositories.semester_repo import SemesterRepo
from repositories.unit_repo import UnitRepo
from repositories.unit_event_submissions_repo import UnitEventSubmissionsRepo
from repositories.unit_event_repo import UnitEventRepo
from schemas.auth import TokenData
from schemas.report import (
    InternalEventCreate,
    InternalEventRead,
    InternalEventUpdate,
    InternalSummary,
    ReportDetail,
    ReportSummary,
    UnitEventSummary,
    ReportPaginationResponse,
)


class ReportService:
    @staticmethod
    async def _get_report_or_404(report_id: PydanticObjectId):
        report = await ReportRepository.get_by_id(report_id)
        if not report:
            app_exception(ErrorCode.REPORT_NOT_FOUND)
        return report

    @staticmethod
    def _get_default_month_year():
        now = datetime.now()
        if now.day <= 20:
            return now.month, now.year
        
        # After 20th, move to next month
        m = now.month + 1
        y = now.year
        if m > 12:
            m = 1
            y += 1
        return m, y

    @staticmethod
    async def _check_deadline_and_status(report: Report):
        # Deadline is 20th of the report's month
        now = datetime.now()
        deadline = datetime(report.year, report.month, 20, 23, 59, 59)
        
        if now > deadline:
            if report.status == "CHUA_NOP":
                report.status = "CHO_DUYET"
                await ReportRepository.save(report)
            return False # Expired
            
        if report.status != "CHUA_NOP":
            return False # Already submitted
            
        return True

    @staticmethod
    async def auto_create_report_for_unit(unit_id: PydanticObjectId, month: int, year: int, semester_id: PydanticObjectId):
        existing = await ReportRepository.get_by_unique(
            unit_id=unit_id,
            month=month,
            year=year,
        )

        if not existing:
            await ReportRepository.create(
                {
                    "unit_id": unit_id,
                    "month": month,
                    "year": year,
                    "semester_id": semester_id,
                    "unit_event_ids": [],
                    "internal_events": [],
                    "status": "CHUA_NOP",
                    "updated_at": datetime.utcnow()
                }
            )

    @staticmethod
    async def auto_create_monthly_reports():
        semester = await SemesterRepo().get_active()
        if not semester:
            return

        now = datetime.now()
        month, year = ReportService._get_default_month_year()

        units = await UnitRepo().list_all()

        for unit in units:
            # 1. Ensure current cycle report exists
            await ReportService.auto_create_report_for_unit(unit.id, month, year, semester.id)

            # 2. Auto-submit any expired reports for this unit
            # Check previous month(s) report
            all_reports = await ReportRepository.get_by_unit(unit.id)
            for r in all_reports:
                deadline = datetime(r.year, r.month, 20, 23, 59, 59)
                if now > deadline and r.status == "CHUA_NOP":
                    r.status = "CHO_DUYET"
                    await ReportRepository.save(r)

    @staticmethod
    async def get_reports_by_unit(
        unit_id: PydanticObjectId,
        month: Optional[int] = None,
        year: Optional[int] = None,
        status_filter: Optional[str] = None,
        skip: int = 0,
        limit: int = 10
    ) -> ReportPaginationResponse:
        # Ensure current cycle report exists
        semester = await SemesterRepo().get_active()
        if semester:
            m, y = ReportService._get_default_month_year()
            await ReportService.auto_create_report_for_unit(unit_id, m, y, semester.id)

        reports, total = await ReportRepository.get_filtered(
            unit_id=unit_id,
            month=month,
            year=year,
            status=status_filter,
            skip=skip,
            limit=limit
        )

        return ReportPaginationResponse(
            items=[
                ReportSummary(
                    id=report.id,
                    unit_id=report.unit_id,
                    month=report.month,
                    semester_id=report.semester_id,
                    year=report.year,
                    status=report.status,
                    updated_at=report.updated_at,
                    total_activities=len(report.unit_event_ids) + len(report.internal_events)
                )
                for report in reports
            ],
            total=total
        )

    @staticmethod
    async def get_all_reports(
        month: Optional[int] = None,
        year: Optional[int] = None,
        unit_id: Optional[PydanticObjectId] = None,
        status_filter: Optional[str] = None,
        skip: int = 0,
        limit: int = 10
    ) -> ReportPaginationResponse:
        # If no filters provided, default to current cycle and exclude drafts
        if month is None and year is None and unit_id is None and status_filter is None:
            month, year = ReportService._get_default_month_year()
        
        reports, total = await ReportRepository.get_filtered(
            month=month,
            year=year,
            unit_id=unit_id,
            status=status_filter,
            skip=skip,
            limit=limit
        )

        return ReportPaginationResponse(
            items=[
                ReportSummary(
                    id=report.id,
                    unit_id=report.unit_id,
                    semester_id=report.semester_id,
                    month=report.month,
                    year=report.year,
                    status=report.status,
                    updated_at=report.updated_at,
                    total_activities=len(report.unit_event_ids) + len(report.internal_events)
                )
                for report in reports 
                if status_filter is not None or report.status != "CHUA_NOP"
            ],
            total=total
        )

    @staticmethod
    async def sync_unit_events(report: Report):
        # 1. Determine period
        if report.month == 1:
            start_date = datetime(report.year - 1, 12, 21)
        else:
            start_date = datetime(report.year, report.month - 1, 21)
        end_date = datetime(report.year, report.month, 20, 23, 59, 59)

        # 2. Get approved submissions
        submissions = await UnitEventSubmissionsRepo().get_approved_by_unit_and_date_range(
            unit_id=report.unit_id,
            start_date=start_date,
            end_date=end_date
        )

        # 3. Update report.unit_event_ids
        event_ids = [s.unitEventId for s in submissions]
        report.unit_event_ids = list(set(event_ids))
        await ReportRepository.save(report)

    @staticmethod
    async def get_report_detail(
        report_id: PydanticObjectId,
        user: TokenData
    ) -> ReportDetail:
        report = await ReportService._get_report_or_404(report_id)

        # Security check: If report is "CHUA_NOP" (Draft),
        # only allow if user has any role at this specific unit
        if report.status == "CHUA_NOP":
            has_local_role = any(str(u_role.unit_id) == str(report.unit_id) for u_role in user.roles)
            if not has_local_role:
                app_exception(
                    ErrorCode.INSUFFICIENT_PERMISSION,
                    extra_detail="Báo cáo nháp chỉ có thể xem bởi nhân sự của đơn vị đó"
                )

        # Auto-sync Type U events
        await ReportService.sync_unit_events(report)
        # Check deadline
        await ReportService._check_deadline_and_status(report)
        unit_events = await UnitEventRepo.get_by_ids(report.unit_event_ids)

        return ReportDetail(
            id=report.id,
            unit_id=report.unit_id,
            month=report.month,
            year=report.year,
            status=report.status,
            note=report.note,
            updated_at=report.updated_at,
            unit_events=[
                UnitEventSummary(
                    id=ue.id, 
                    title=ue.title,
                    type=ue.type,
                    created_at=ue.created_at
                )
                for ue in unit_events
            ],
            internal_events=[
                InternalEventRead(**event.model_dump())
                for event in report.internal_events
            ],
        )

    @staticmethod
    async def add_internal_event(
        report_id: PydanticObjectId,
        data: InternalEventCreate,
    ) -> InternalEventRead:
        report = await ReportService._get_report_or_404(report_id)
        
        if not await ReportService._check_deadline_and_status(report):
            app_exception(ErrorCode.REPORT_LOCKED)

        event = InternalEvent(**data.model_dump())
        report.internal_events.append(event)

        await ReportRepository.save(report)

        return InternalEventRead(**event.model_dump())

    @staticmethod
    async def update_internal_event(
        report_id: PydanticObjectId,
        event_id: PydanticObjectId,
        data: InternalEventUpdate,
    ) -> InternalEventRead:
        report = await ReportService._get_report_or_404(report_id)

        if not await ReportService._check_deadline_and_status(report):
            app_exception(ErrorCode.REPORT_LOCKED)

        event = next(
            (item for item in report.internal_events if str(item.id) == str(event_id)),
            None,
        )

        if not event:
            app_exception(ErrorCode.EVENT_NOT_FOUND)

        update_data = data.model_dump(exclude_unset=True)

        for field, value in update_data.items():
            setattr(event, field, value)

        await ReportRepository.save(report)

        return InternalEventRead(**event.model_dump())

    @staticmethod
    async def delete_internal_event(
        report_id: PydanticObjectId,
        event_id: PydanticObjectId,
    ):
        report = await ReportService._get_report_or_404(report_id)

        if not await ReportService._check_deadline_and_status(report):
            app_exception(ErrorCode.REPORT_LOCKED)

        report.internal_events = [
            item for item in report.internal_events if str(item.id) != str(event_id)
        ]

        await ReportRepository.save(report)

    @staticmethod
    async def save(report: Report):
        await ReportRepository.save(report)

    @staticmethod
    async def submit_report(report_id: PydanticObjectId):
        report = await ReportService._get_report_or_404(report_id)
        
        if report.status == "CHO_DUYET" or report.status == "DA_DUYET":
            return report

        report.status = "CHO_DUYET"
        report.updated_at = datetime.utcnow()
        await ReportRepository.save(report)
        return report

    @staticmethod
    async def update_report_status(report_id: PydanticObjectId, status: str, note: Optional[str] = None):
        report = await ReportService._get_report_or_404(report_id)
        report.status = status
        if note is not None:
            report.note = note
        report.updated_at = datetime.utcnow()
        await ReportRepository.save(report)
        return report

    @staticmethod
    async def export_summary_excel(
        month: Optional[int] = None,
        year: Optional[int] = None,
        unit_id: Optional[PydanticObjectId] = None,
        status_filter: Optional[str] = None
    ):
        # We fetch a large number for export, effectively disabling pagination for excel
        reports, _ = await ReportRepository.get_filtered(
            month=month,
            year=year,
            unit_id=unit_id,
            status=status_filter,
            limit=5000 
        )
        
        data = []
        for r in reports:
            # Skip drafts in summary unless explicitly filtered
            if status_filter is None and r.status == "CHUA_NOP":
                continue
                
            unit = await UnitRepo().get_by_id(r.unit_id)
            data.append({
                "Đơn vị": unit.name if unit else str(r.unit_id),
                "Tháng": r.month,
                "Năm": r.year,
                "Số hoạt động": len(r.unit_event_ids) + len(r.internal_events),
                "Trạng thái": r.status,
                "Ngày cập nhật": r.updated_at.strftime("%d/%m/%Y %H:%M") if r.updated_at else "N/A",
                "Ghi chú": r.note or ""
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Tong_hop_bao_cao')
            
            # Formatting
            workbook = writer.book
            worksheet = writer.sheets['Tong_hop_bao_cao']
            
            # Header style
            from openpyxl.styles import Font, Alignment, PatternFill
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column].width = max_length + 5

        output.seek(0)
        return output

    @staticmethod
    async def export_detailed_excel(report_id: PydanticObjectId, user: TokenData):
        report = await ReportService._get_report_or_404(report_id)
        
        # Security check: Managers can export anything. Staff only their own unit's.
        is_manager = any("ADMIN" in r.roles or "MANAGER" in r.roles for r in user.roles)
        if not is_manager:
            is_staff_of_unit = any(str(r.unit_id) == str(report.unit_id) for r in user.roles)
            if not is_staff_of_unit:
                app_exception(ErrorCode.INSUFFICIENT_PERMISSION)

        unit = await UnitRepo().get_by_id(report.unit_id)
        unit_name = unit.name if unit else "N/A"

        unit_events = await UnitEventRepo.get_by_ids(report.unit_event_ids)
        
        data = []
        # Unit Events (Assigned)
        for ue in unit_events:
            data.append({
                "Tên sự kiện/hoạt động": ue.title,
                "Loại hình": "Được phân công (Hệ thống)",
                "Thời gian diễn ra": ue.created_at.strftime("%d/%m/%Y %H:%M") if ue.created_at else "N/A",
                "Địa điểm": "Dữ liệu hệ thống",
                "Số người tham gia": "Theo đăng ký",
                "Ghi chú": "Hệ thống tự động đồng bộ"
            })
        
        # Internal Events
        for ie in report.internal_events:
            data.append({
                "Tên sự kiện/hoạt động": ie.title,
                "Loại hình": "Hoạt động nội bộ (Tự quản)",
                "Thời gian diễn ra": ie.event_date.strftime("%d/%m/%Y") if ie.event_date else "N/A",
                "Địa điểm": ie.location or "N/A",
                "Số người tham gia": ie.participant_count or 0,
                "Ghi chú": ie.description or ""
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = f'Bao_cao_{report.month}_{report.year}'
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Formatting
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            from openpyxl.styles import Font, Alignment, PatternFill
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto column width
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column].width = min(max_length + 5, 50) # Cap at 50

        output.seek(0)
        return output
